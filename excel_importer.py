#!/usr/bin/env python3
"""
Excel Importer for Knowledge Base Content
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from models import KnowledgeArticle, SolutionStep, DiagnosticQuestion
from utils import DataValidator, DataConverter
from csv_importer import ImportResult


class ExcelImporter:
    """Excel file importer for knowledge base content."""
    
    def __init__(self, es_manager=None):
        """Initialize the Excel importer."""
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            raise ImportError("Neither openpyxl nor pandas is available. Please install one of them.")
        
        self.es_manager = es_manager
        self.validator = DataValidator()
        self.converter = DataConverter()
        
        self.required_columns = [
            'title', 'category', 'subcategory', 'content', 
            'keywords', 'symptoms', 'difficulty', 'estimated_time'
        ]
        
        self.import_stats = {
            'total_processed': 0,
            'successful': 0,
            'failed': 0,
            'errors': [],
            'warnings': []
        }
    
    def import_from_excel(self, file_path: str, preview_mode: bool = False) -> ImportResult:
        """Import content from an Excel file."""
        start_time = datetime.now()
        self._reset_stats()
        
        try:
            if not self._file_exists(file_path):
                raise FileNotFoundError(f"Excel file not found: {file_path}")
            
            logging.info(f"Starting Excel import from: {file_path}")
            
            # Use pandas if available, otherwise openpyxl
            if PANDAS_AVAILABLE:
                articles_data = self._import_with_pandas(file_path)
            else:
                articles_data = self._import_with_openpyxl(file_path)
            
            # Validate and convert articles
            valid_articles = self._validate_articles(articles_data)
            
            # Import to Elasticsearch if not in preview mode
            if not preview_mode and self.es_manager and valid_articles:
                self._bulk_import(valid_articles)
            
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=self.import_stats['failed'] == 0,
                total_records=self.import_stats['total_processed'],
                successful_imports=self.import_stats['successful'],
                failed_imports=self.import_stats['failed'],
                errors=self.import_stats['errors'],
                warnings=self.import_stats['warnings'],
                processing_time=processing_time
            )
            
        except Exception as e:
            logging.error(f"Excel import failed: {e}")
            processing_time = (datetime.now() - start_time).total_seconds()
            return ImportResult(
                success=False,
                total_records=0,
                successful_imports=0,
                failed_imports=1,
                errors=[{"type": "import_error", "message": str(e)}],
                warnings=[],
                processing_time=processing_time
            )
    
    def _import_with_pandas(self, file_path: str) -> List[Dict[str, Any]]:
        """Import using pandas."""
        try:
            # Read the main articles sheet
            df = pd.read_excel(file_path, sheet_name='Articles')
            
            # Convert to list of dictionaries
            articles_data = []
            for index, row in df.iterrows():
                row_data = {}
                for column in df.columns:
                    value = row[column]
                    if pd.notna(value):
                        row_data[str(column).lower().replace(' ', '_')] = str(value).strip()
                
                if row_data:  # Only add if we have data
                    row_data['_row_number'] = index + 2  # +2 because Excel is 1-based and we have header
                    articles_data.append(row_data)
            
            # Try to read categories sheet if available
            try:
                categories_df = pd.read_excel(file_path, sheet_name='Categories')
                logging.info(f"Found {len(categories_df)} categories")
            except:
                pass
            
            return articles_data
            
        except Exception as e:
            logging.warning(f"Pandas import failed, falling back to openpyxl: {e}")
            return self._import_with_openpyxl(file_path)
    
    def _import_with_openpyxl(self, file_path: str) -> List[Dict[str, Any]]:
        """Import using openpyxl."""
        workbook = load_workbook(file_path, data_only=True)
        
        # Process main articles sheet
        if 'Articles' in workbook.sheetnames:
            articles_sheet = workbook['Articles']
        else:
            # Use first sheet if 'Articles' not found
            articles_sheet = workbook.active
        
        articles_data = self._process_articles_sheet(articles_sheet)
        
        # Process categories sheet if available
        if 'Categories' in workbook.sheetnames:
            categories_sheet = workbook['Categories']
            categories_data = self._process_categories_sheet(categories_sheet)
            logging.info(f"Found {len(categories_data)} categories")
        
        return articles_data
    
    def _process_articles_sheet(self, sheet) -> List[Dict[str, Any]]:
        """Process the articles worksheet."""
        articles = []
        
        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip().lower().replace(' ', '_'))
        
        # Validate headers
        missing_required = [col for col in self.required_columns if col not in headers]
        if missing_required:
            raise ValueError(f"Missing required columns: {missing_required}")
        
        # Process data rows
        for row_num in range(2, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_num, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data[header] = str(cell_value).strip()
                    has_data = True
            
            if has_data:
                try:
                    article_data = self._process_excel_row(row_data, row_num)
                    if article_data:
                        articles.append(article_data)
                except Exception as e:
                    self._record_error(row_num, "row_processing", str(e))
        
        return articles
    
    def _process_categories_sheet(self, sheet) -> List[Dict[str, Any]]:
        """Process the categories worksheet."""
        categories = []
        
        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip().lower().replace(' ', '_'))
        
        # Process data rows
        for row_num in range(2, sheet.max_row + 1):
            category_data = {}
            has_data = False
            
            for col_num, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    category_data[header] = str(cell_value).strip()
                    has_data = True
            
            if has_data:
                categories.append(category_data)
        
        return categories
    
    def _process_excel_row(self, row_data: Dict[str, str], row_num: int) -> Optional[Dict[str, Any]]:
        """Process a single Excel row into article data."""
        try:
            # Basic data extraction
            article_data = {
                'title': row_data.get('title', '').strip(),
                'category': row_data.get('category', '').strip(),
                'subcategory': row_data.get('subcategory', '').strip(),
                'content': row_data.get('content', '').strip(),
                'keywords': self._parse_keywords(row_data.get('keywords', '')),
                'symptoms': self._parse_symptoms(row_data.get('symptoms', '')),
                'difficulty_level': row_data.get('difficulty', 'medium').strip().lower(),
                'estimated_time_minutes': self._parse_int(row_data.get('estimated_time', '0')),
                'success_rate': self._parse_float(row_data.get('success_rate', '0.8')),
                'solution_steps': self._parse_solution_steps(row_data.get('solution_steps', '')),
                'diagnostic_questions': self._parse_diagnostic_questions(row_data.get('diagnostic_questions', '')),
                'is_active': True,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                '_row_number': row_num
            }
            
            # Validate required fields
            if not article_data['title'] or not article_data['content']:
                raise ValueError("Title and content are required")
            
            return article_data
            
        except Exception as e:
            self._record_error(row_num, "row_processing", str(e))
            return None
    
    def _parse_keywords(self, keywords_str: str) -> List[str]:
        """Parse keywords from comma-separated string."""
        if not keywords_str:
            return []
        return [kw.strip() for kw in keywords_str.split(',') if kw.strip()]
    
    def _parse_symptoms(self, symptoms_str: str) -> List[str]:
        """Parse symptoms from comma-separated string."""
        if not symptoms_str:
            return []
        return [symptom.strip() for symptom in symptoms_str.split(',') if symptom.strip()]
    
    def _parse_solution_steps(self, steps_str: str) -> List[Dict[str, Any]]:
        """Parse solution steps from string."""
        if not steps_str:
            return []
        
        steps = []
        lines = steps_str.strip().split('\n')
        step_num = 1
        
        for line in lines:
            line = line.strip()
            if line:
                steps.append({
                    'order': step_num,
                    'title': f"Step {step_num}",
                    'content': line,
                    'step_type': 'instruction'
                })
                step_num += 1
        
        return steps
    
    def _parse_diagnostic_questions(self, questions_str: str) -> List[Dict[str, Any]]:
        """Parse diagnostic questions from string."""
        if not questions_str:
            return []
        
        questions = []
        lines = questions_str.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if line:
                questions.append({
                    'question': line,
                    'question_type': 'text',
                    'required': False
                })
        
        return questions
    
    def _parse_int(self, value: str, default: int = 0) -> int:
        """Parse integer value with default."""
        try:
            return int(value.strip())
        except (ValueError, AttributeError):
            return default
    
    def _parse_float(self, value: str, default: float = 0.0) -> float:
        """Parse float value with default."""
        try:
            return float(value.strip())
        except (ValueError, AttributeError):
            return default
    
    def _validate_articles(self, articles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Validate articles and return valid ones."""
        valid_articles = []
        
        for article_data in articles:
            try:
                self.import_stats['total_processed'] += 1
                
                # Validate article data
                is_valid, errors = self.validator.validate_article_data(article_data)
                if is_valid:
                    # Convert to Elasticsearch format
                    es_doc = self.converter.article_to_elasticsearch(article_data)
                    valid_articles.append(es_doc)
                    self.import_stats['successful'] += 1
                else:
                    for error in errors:
                        self._record_error(
                            article_data.get('_row_number'), "validation", error
                        )
                    self.import_stats['failed'] += 1
                    
            except Exception as e:
                self._record_error(
                    article_data.get('_row_number'), "conversion", str(e)
                )
                self.import_stats['failed'] += 1
        
        return valid_articles
    
    def _bulk_import(self, articles: List[Dict[str, Any]]):
        """Perform bulk import to Elasticsearch."""
        try:
            bulk_result = self.es_manager.bulk_index_articles(articles)
            logging.info(f"Bulk import result: {bulk_result}")
        except Exception as e:
            logging.error(f"Bulk import failed: {e}")
            self._record_error(None, "bulk_import", str(e))
    
    def _reset_stats(self):
        """Reset import statistics."""
        self.import_stats = {
            'total_processed': 0,
            'successful': 0,
            'failed': 0,
            'errors': [],
            'warnings': []
        }
    
    def _file_exists(self, file_path: str) -> bool:
        """Check if file exists."""
        return Path(file_path).exists()
    
    def _record_error(self, row_number: Optional[int], error_type: str, message: str):
        """Record an error."""
        error_record = {
            'type': error_type,
            'message': message,
            'row_number': row_number
        }
        self.import_stats['errors'].append(error_record)


# Configure logging
logging.basicConfig(level=logging.INFO)
