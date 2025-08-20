#!/usr/bin/env python3
"""
Knowledge Base Content Import System

This module provides comprehensive import functionality for helpdesk knowledge base content
from various formats including CSV, JSON, and Excel files.
"""

import csv
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models import KnowledgeArticle, SolutionStep, DiagnosticQuestion, DifficultyLevel
from utils import DataValidator, DataConverter, IDGenerator
from config_manager import get_config_manager


# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """Result of an import operation."""
    success: bool
    total_records: int
    successful_imports: int
    failed_imports: int
    errors: List[Dict[str, Any]]
    warnings: List[str]
    processing_time: float


@dataclass
class ValidationError:
    """Validation error details."""
    row_number: Optional[int]
    field_name: str
    error_message: str
    severity: str = "error"  # error, warning, info


class ContentImporter:
    """Main content importer class."""
    
    def __init__(self, es_manager=None):
        """Initialize the importer."""
        self.es_manager = es_manager
        self.config = get_config_manager()
        self.validator = DataValidator()
        self.converter = DataConverter()
        
        # Track import statistics
        self.import_stats = {
            'total_processed': 0,
            'successful': 0,
            'failed': 0,
            'skipped': 0,
            'errors': [],
            'warnings': []
        }
    
    def reset_stats(self):
        """Reset import statistics."""
        self.import_stats = {
            'total_processed': 0,
            'successful': 0,
            'failed': 0,
            'skipped': 0,
            'errors': [],
            'warnings': []
        }


class CSVImporter(ContentImporter):
    """CSV file importer."""
    
    def __init__(self, es_manager=None):
        super().__init__(es_manager)
        self.required_columns = [
            'title', 'category', 'subcategory', 'content', 
            'keywords', 'symptoms', 'difficulty', 'estimated_time'
        ]
        self.optional_columns = ['solution_steps', 'diagnostic_questions', 'success_rate']
    
    def import_from_csv(self, file_path: str, preview_mode: bool = False) -> ImportResult:
        """Import content from a CSV file."""
        start_time = datetime.now()
        self.reset_stats()
        
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"CSV file not found: {file_path}")
            
            logger.info(f"Starting CSV import from: {file_path}")
            
            with open(file_path, 'r', encoding='utf-8') as file:
                # Detect CSV dialect
                sample = file.read(1024)
                file.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                
                reader = csv.DictReader(file, dialect=dialect)
                
                # Validate headers
                header_validation = self._validate_csv_headers(reader.fieldnames)
                if not header_validation['valid']:
                    raise ValueError(f"Invalid CSV headers: {header_validation['errors']}")
                
                # Process rows
                articles = []
                for row_num, row in enumerate(reader, start=2):  # Start at 2 (1-based + header)
                    try:
                        article_data = self._process_csv_row(row, row_num)
                        if article_data:
                            articles.append(article_data)
                            self.import_stats['total_processed'] += 1
                    except Exception as e:
                        self._record_error(row_num, "row_processing", str(e))
                        self.import_stats['failed'] += 1
                
                # Validate and convert articles
                valid_articles = []
                for article_data in articles:
                    try:
                        # Validate article data
                        is_valid, errors = self.validator.validate_article_data(article_data)
                        if is_valid:
                            # Convert to Elasticsearch format
                            es_doc = self.converter.article_to_elasticsearch(article_data)
                            valid_articles.append(es_doc)
                            self.import_stats['successful'] += 1
                        else:
                            for error in errors:
                                self._record_error(
                                    None, "validation", error, 
                                    row_number=article_data.get('_row_number')
                                )
                            self.import_stats['failed'] += 1
                    except Exception as e:
                        self._record_error(
                            None, "conversion", str(e),
                            row_number=article_data.get('_row_number')
                        )
                        self.import_stats['failed'] += 1
                
                # Import to Elasticsearch if not in preview mode
                if not preview_mode and self.es_manager and valid_articles:
                    try:
                        bulk_result = self.es_manager.bulk_index_articles(valid_articles)
                        logger.info(f"Bulk import result: {bulk_result}")
                    except Exception as e:
                        logger.error(f"Bulk import failed: {e}")
                        self._record_error(None, "bulk_import", str(e))
                
                processing_time = (datetime.now() - start_time).total_seconds()
                
                return ImportResult(
                    success=self.import_stats['failed'] == 0,
                    total_records=self.import_stats['total_processed'],
                    successful_imports=self.import_stats['successful'],
                    failed_imports=self.import_stats['failed'],
                    errors=self.import_stats['errors'],
                    warnings=self.import_stats['warnings'],
                    processing_time=processing_time
                )
                
        except Exception as e:
            logger.error(f"CSV import failed: {e}")
            processing_time = (datetime.now() - start_time).total_seconds()
            return ImportResult(
                success=False,
                total_records=0,
                successful_imports=0,
                failed_imports=1,
                errors=[{"type": "import_error", "message": str(e)}],
                warnings=[],
                processing_time=processing_time
            )
    
    def _validate_csv_headers(self, fieldnames: List[str]) -> Dict[str, Any]:
        """Validate CSV headers against required columns."""
        if not fieldnames:
            return {'valid': False, 'errors': ['No headers found']}
        
        missing_required = [col for col in self.required_columns if col not in fieldnames]
        if missing_required:
            return {
                'valid': False, 
                'errors': [f"Missing required columns: {missing_required}"]
            }
        
        return {'valid': True, 'errors': []}
    
    def _process_csv_row(self, row: Dict[str, str], row_num: int) -> Optional[Dict[str, Any]]:
        """Process a single CSV row into article data."""
        try:
            # Basic data extraction
            article_data = {
                'title': row.get('title', '').strip(),
                'category': row.get('category', '').strip(),
                'subcategory': row.get('subcategory', '').strip(),
                'content': row.get('content', '').strip(),
                'keywords': self._parse_keywords(row.get('keywords', '')),
                'symptoms': self._parse_symptoms(row.get('symptoms', '')),
                'difficulty_level': row.get('difficulty', 'medium').strip().lower(),
                'estimated_time_minutes': self._parse_int(row.get('estimated_time', '0')),
                'success_rate': self._parse_float(row.get('success_rate', '0.8')),
                'solution_steps': self._parse_solution_steps(row.get('solution_steps', '')),
                'diagnostic_questions': self._parse_diagnostic_questions(row.get('diagnostic_questions', '')),
                'is_active': True,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                '_row_number': row_num
            }
            
            # Validate required fields
            if not article_data['title'] or not article_data['content']:
                raise ValueError("Title and content are required")
            
            return article_data
            
        except Exception as e:
            self._record_error(row_num, "row_processing", str(e))
            return None
    
    def _parse_keywords(self, keywords_str: str) -> List[str]:
        """Parse keywords from comma-separated string."""
        if not keywords_str:
            return []
        return [kw.strip() for kw in keywords_str.split(',') if kw.strip()]
    
    def _parse_symptoms(self, symptoms_str: str) -> List[str]:
        """Parse symptoms from comma-separated string."""
        if not symptoms_str:
            return []
        return [symptom.strip() for symptom in symptoms_str.split(',') if symptom.strip()]
    
    def _parse_solution_steps(self, steps_str: str) -> List[Dict[str, Any]]:
        """Parse solution steps from string or JSON."""
        if not steps_str:
            return []
        
        try:
            # Try to parse as JSON first
            if steps_str.strip().startswith('['):
                steps_data = json.loads(steps_str)
                if isinstance(steps_data, list):
                    return steps_data
        except json.JSONDecodeError:
            pass
        
        # Parse as numbered list
        steps = []
        lines = steps_str.strip().split('\n')
        step_num = 1
        
        for line in lines:
            line = line.strip()
            if line:
                # Remove numbering if present
                if line[0].isdigit() and '.' in line[:3]:
                    line = line.split('.', 1)[1].strip()
                
                if line:
                    steps.append({
                        'order': step_num,
                        'title': f"Step {step_num}",
                        'content': line,
                        'step_type': 'instruction'
                    })
                    step_num += 1
        
        return steps
    
    def _parse_diagnostic_questions(self, questions_str: str) -> List[Dict[str, Any]]:
        """Parse diagnostic questions from string or JSON."""
        if not questions_str:
            return []
        
        try:
            # Try to parse as JSON first
            if questions_str.strip().startswith('['):
                questions_data = json.loads(questions_str)
                if isinstance(questions_data, list):
                    return questions_data
        except json.JSONDecodeError:
            pass
        
        # Parse as simple questions
        questions = []
        lines = questions_str.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if line:
                questions.append({
                    'question': line,
                    'question_type': 'text',
                    'required': False
                })
        
        return questions
    
    def _parse_int(self, value: str, default: int = 0) -> int:
        """Parse integer value with default."""
        try:
            return int(value.strip())
        except (ValueError, AttributeError):
            return default
    
    def _parse_float(self, value: str, default: float = 0.0) -> float:
        """Parse float value with default."""
        try:
            return float(value.strip())
        except (ValueError, AttributeError):
            return default
    
    def _record_error(self, row_number: Optional[int], error_type: str, 
                     message: str, severity: str = "error"):
        """Record an error or warning."""
        error_record = {
            'type': error_type,
            'message': message,
            'severity': severity
        }
        if row_number:
            error_record['row_number'] = row_number
        
        if severity == "error":
            self.import_stats['errors'].append(error_record)
        else:
            self.import_stats['warnings'].append(error_record)


class JSONImporter(ContentImporter):
    """JSON file importer."""
    
    def __init__(self, es_manager=None):
        super().__init__(es_manager)
    
    def import_from_json(self, file_path: str, preview_mode: bool = False, 
                        update_existing: bool = False) -> ImportResult:
        """Import content from a JSON file."""
        start_time = datetime.now()
        self.reset_stats()
        
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"JSON file not found: {file_path}")
            
            logger.info(f"Starting JSON import from: {file_path}")
            
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            # Handle different JSON formats
            if isinstance(data, list):
                articles_data = data
            elif isinstance(data, dict) and 'articles' in data:
                articles_data = data['articles']
            else:
                raise ValueError("Invalid JSON format: expected array of articles or object with 'articles' key")
            
            valid_articles = []
            
            for i, article_data in enumerate(articles_data):
                try:
                    self.import_stats['total_processed'] += 1
                    
                    # Validate article data
                    is_valid, errors = self.validator.validate_article_data(article_data)
                    if is_valid:
                        # Convert to Elasticsearch format
                        es_doc = self.converter.article_to_elasticsearch(article_data)
                        valid_articles.append(es_doc)
                        self.import_stats['successful'] += 1
                    else:
                        for error in errors:
                            self._record_error(i + 1, "validation", error)
                        self.import_stats['failed'] += 1
                        
                except Exception as e:
                    self._record_error(i + 1, "processing", str(e))
                    self.import_stats['failed'] += 1
            
            # Import to Elasticsearch if not in preview mode
            if not preview_mode and self.es_manager and valid_articles:
                try:
                    if update_existing:
                        # Handle updates
                        for article in valid_articles:
                            if 'article_id' in article:
                                self.es_manager.update_article(article['article_id'], article)
                            else:
                                self.es_manager.index_article(article)
                    else:
                        # Bulk import
                        bulk_result = self.es_manager.bulk_index_articles(valid_articles)
                        logger.info(f"Bulk import result: {bulk_result}")
                        
                except Exception as e:
                    logger.error(f"Import failed: {e}")
                    self._record_error(None, "import", str(e))
            
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=self.import_stats['failed'] == 0,
                total_records=self.import_stats['total_processed'],
                successful_imports=self.import_stats['successful'],
                failed_imports=self.import_stats['failed'],
                errors=self.import_stats['errors'],
                warnings=self.import_stats['warnings'],
                processing_time=processing_time
            )
            
        except Exception as e:
            logger.error(f"JSON import failed: {e}")
            processing_time = (datetime.now() - start_time).total_seconds()
            return ImportResult(
                success=False,
                total_records=0,
                successful_imports=0,
                failed_imports=1,
                errors=[{"type": "import_error", "message": str(e)}],
                warnings=[],
                processing_time=processing_time
            )
    
    def _record_error(self, index: Optional[int], error_type: str, message: str):
        """Record an error."""
        error_record = {
            'type': error_type,
            'message': message,
            'index': index
        }
        self.import_stats['errors'].append(error_record)


class ExcelImporter(ContentImporter):
    """Excel file importer using openpyxl and pandas."""
    
    def __init__(self, es_manager=None):
        super().__init__(es_manager)
        self.required_columns = [
            'title', 'category', 'subcategory', 'content', 
            'keywords', 'symptoms', 'difficulty', 'estimated_time'
        ]
    
    def import_from_excel(self, file_path: str, preview_mode: bool = False) -> ImportResult:
        """Import content from an Excel file."""
        start_time = datetime.now()
        self.reset_stats()
        
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Excel file not found: {file_path}")
            
            logger.info(f"Starting Excel import from: {file_path}")
            
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            # Process main articles sheet
            if 'Articles' in workbook.sheetnames:
                articles_sheet = workbook['Articles']
                articles_data = self._process_articles_sheet(articles_sheet)
            else:
                # Use first sheet if 'Articles' not found
                articles_data = self._process_articles_sheet(workbook.active)
            
            # Process categories sheet if available
            if 'Categories' in workbook.sheetnames:
                categories_sheet = workbook['Categories']
                categories_data = self._process_categories_sheet(categories_sheet)
                logger.info(f"Found {len(categories_data)} categories")
            
            # Validate and convert articles
            valid_articles = []
            for i, article_data in enumerate(articles_data):
                try:
                    self.import_stats['total_processed'] += 1
                    
                    # Validate article data
                    is_valid, errors = self.validator.validate_article_data(article_data)
                    if is_valid:
                        # Convert to Elasticsearch format
                        es_doc = self.converter.article_to_elasticsearch(article_data)
                        valid_articles.append(es_doc)
                        self.import_stats['successful'] += 1
                    else:
                        for error in errors:
                            self._record_error(i + 1, "validation", error)
                        self.import_stats['failed'] += 1
                        
                except Exception as e:
                    self._record_error(i + 1, "processing", str(e))
                    self.import_stats['failed'] += 1
            
            # Import to Elasticsearch if not in preview mode
            if not preview_mode and self.es_manager and valid_articles:
                try:
                    bulk_result = self.es_manager.bulk_index_articles(valid_articles)
                    logger.info(f"Bulk import result: {bulk_result}")
                except Exception as e:
                    logger.error(f"Import failed: {e}")
                    self._record_error(None, "import", str(e))
            
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=self.import_stats['failed'] == 0,
                total_records=self.import_stats['total_processed'],
                successful_imports=self.import_stats['successful'],
                failed_imports=self.import_stats['failed'],
                errors=self.import_stats['errors'],
                warnings=self.import_stats['warnings'],
                processing_time=processing_time
            )
            
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            processing_time = (datetime.now() - start_time).total_seconds()
            return ImportResult(
                success=False,
                total_records=0,
                successful_imports=0,
                failed_imports=1,
                errors=[{"type": "import_error", "message": str(e)}],
                warnings=[],
                processing_time=processing_time
            )
    
    def _process_articles_sheet(self, sheet) -> List[Dict[str, Any]]:
        """Process the articles worksheet."""
        articles = []
        
        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip().lower())
        
        # Validate headers
        missing_required = [col for col in self.required_columns if col not in headers]
        if missing_required:
            raise ValueError(f"Missing required columns: {missing_required}")
        
        # Process data rows
        for row_num in range(2, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_num, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data[header] = str(cell_value).strip()
                    has_data = True
            
            if has_data:
                try:
                    article_data = self._process_excel_row(row_data, row_num)
                    if article_data:
                        articles.append(article_data)
                except Exception as e:
                    self._record_error(row_num, "row_processing", str(e))
        
        return articles
    
    def _process_categories_sheet(self, sheet) -> List[Dict[str, Any]]:
        """Process the categories worksheet."""
        categories = []
        
        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip().lower())
        
        # Process data rows
        for row_num in range(2, sheet.max_row + 1):
            category_data = {}
            has_data = False
            
            for col_num, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    category_data[header] = str(cell_value).strip()
                    has_data = True
            
            if has_data:
                categories.append(category_data)
        
        return categories
    
    def _process_excel_row(self, row_data: Dict[str, str], row_num: int) -> Optional[Dict[str, Any]]:
        """Process a single Excel row into article data."""
        try:
            # Basic data extraction
            article_data = {
                'title': row_data.get('title', '').strip(),
                'category': row_data.get('category', '').strip(),
                'subcategory': row_data.get('subcategory', '').strip(),
                'content': row_data.get('content', '').strip(),
                'keywords': self._parse_keywords(row_data.get('keywords', '')),
                'symptoms': self._parse_symptoms(row_data.get('symptoms', '')),
                'difficulty_level': row_data.get('difficulty', 'medium').strip().lower(),
                'estimated_time_minutes': self._parse_int(row_data.get('estimated_time', '0')),
                'success_rate': self._parse_float(row_data.get('success_rate', '0.8')),
                'solution_steps': self._parse_solution_steps(row_data.get('solution_steps', '')),
                'diagnostic_questions': self._parse_diagnostic_questions(row_data.get('diagnostic_questions', '')),
                'is_active': True,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                '_row_number': row_num
            }
            
            # Validate required fields
            if not article_data['title'] or not article_data['content']:
                raise ValueError("Title and content are required")
            
            return article_data
            
        except Exception as e:
            self._record_error(row_num, "row_processing", str(e))
            return None
    
    def _parse_keywords(self, keywords_str: str) -> List[str]:
        """Parse keywords from comma-separated string."""
        if not keywords_str:
            return []
        return [kw.strip() for kw in keywords_str.split(',') if kw.strip()]
    
    def _parse_symptoms(self, symptoms_str: str) -> List[str]:
        """Parse symptoms from comma-separated string."""
        if not symptoms_str:
            return []
        return [symptom.strip() for symptom in symptoms_str.split(',') if symptom.strip()]
    
    def _parse_solution_steps(self, steps_str: str) -> List[Dict[str, Any]]:
        """Parse solution steps from string."""
        if not steps_str:
            return []
        
        steps = []
        lines = steps_str.strip().split('\n')
        step_num = 1
        
        for line in lines:
            line = line.strip()
            if line:
                steps.append({
                    'order': step_num,
                    'title': f"Step {step_num}",
                    'content': line,
                    'step_type': 'instruction'
                })
                step_num += 1
        
        return steps
    
    def _parse_diagnostic_questions(self, questions_str: str) -> List[Dict[str, Any]]:
        """Parse diagnostic questions from string."""
        if not questions_str:
            return []
        
        questions = []
        lines = questions_str.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if line:
                questions.append({
                    'question': line,
                    'question_type': 'text',
                    'required': False
                })
        
        return questions
    
    def _parse_int(self, value: str, default: int = 0) -> int:
        """Parse integer value with default."""
        try:
            return int(value.strip())
        except (ValueError, AttributeError):
            return default
    
    def _parse_float(self, value: str, default: float = 0.0) -> float:
        """Parse float value with default."""
        try:
            return float(value.strip())
        except (ValueError, AttributeError):
            return default
    
    def _record_error(self, row_number: Optional[int], error_type: str, message: str):
        """Record an error."""
        error_record = {
            'type': error_type,
            'message': message,
            'row_number': row_number
        }
        self.import_stats['errors'].append(error_record)


class ContentValidator:
    """Content validation and quality checking."""
    
    def __init__(self):
        self.config = get_config_manager()
    
    def validate_article(self, article_data: Dict[str, Any]) -> Tuple[bool, List[ValidationError]]:
        """Validate a single article."""
        errors = []
        
        # Check required fields
        required_fields = ['title', 'content', 'category']
        for field in required_fields:
            if not article_data.get(field):
                errors.append(ValidationError(
                    row_number=article_data.get('_row_number'),
                    field_name=field,
                    error_message=f"Required field '{field}' is missing",
                    severity="error"
                ))
        
        # Validate data types
        if 'estimated_time_minutes' in article_data:
            try:
                time_value = int(article_data['estimated_time_minutes'])
                if time_value < 0:
                    errors.append(ValidationError(
                        row_number=article_data.get('_row_number'),
                        field_name='estimated_time_minutes',
                        error_message="Estimated time must be positive",
                        severity="error"
                    ))
            except (ValueError, TypeError):
                errors.append(ValidationError(
                    row_number=article_data.get('_row_number'),
                    field_name='estimated_time_minutes',
                    error_message="Estimated time must be a valid integer",
                    severity="error"
                ))
        
        if 'success_rate' in article_data:
            try:
                rate_value = float(article_data['success_rate'])
                if not 0.0 <= rate_value <= 1.0:
                    errors.append(ValidationError(
                        row_number=article_data.get('_row_number'),
                        field_name='success_rate',
                        error_message="Success rate must be between 0.0 and 1.0",
                        severity="error"
                    ))
            except (ValueError, TypeError):
                errors.append(ValidationError(
                    row_number=article_data.get('_row_number'),
                    field_name='success_rate',
                    error_message="Success rate must be a valid number",
                    severity="error"
                ))
        
        # Validate difficulty level
        if 'difficulty_level' in article_data:
            difficulty = article_data['difficulty_level'].lower()
            valid_difficulties = ['easy', 'medium', 'hard']
            if difficulty not in valid_difficulties:
                errors.append(ValidationError(
                    row_number=article_data.get('_row_number'),
                    field_name='difficulty_level',
                    error_message=f"Difficulty must be one of: {valid_difficulties}",
                    severity="error"
                ))
        
        # Check content length
        if 'content' in article_data:
            content = article_data['content']
            if len(content) < 10:
                errors.append(ValidationError(
                    row_number=article_data.get('_row_number'),
                    field_name='content',
                    error_message="Content is too short (minimum 10 characters)",
                    severity="warning"
                ))
        
        # Check for duplicate titles (warning)
        # This would require checking against existing data
        
        return len([e for e in errors if e.severity == "error"]) == 0, errors
    
    def check_category_consistency(self, articles: List[Dict[str, Any]]) -> List[str]:
        """Check for category consistency across articles."""
        warnings = []
        
        # Collect all categories and subcategories
        categories = set()
        subcategories = set()
        
        for article in articles:
            if 'category' in article:
                categories.add(article['category'])
            if 'subcategory' in article:
                subcategories.add(article['subcategory'])
        
        # Check for orphaned subcategories
        orphaned_subcategories = []
        for article in articles:
            if 'subcategory' in article and 'category' in article:
                if article['subcategory'] and not article['category']:
                    orphaned_subcategories.append(article['subcategory'])
        
        if orphaned_subcategories:
            warnings.append(f"Found subcategories without parent categories: {orphaned_subcategories}")
        
        return warnings


def create_sample_csv_template(file_path: str = "sample_articles.csv"):
    """Create a sample CSV template for import."""
    sample_data = [
        {
            'title': 'How to Reset Email Password',
            'category': 'Email',
            'subcategory': 'Password Management',
            'content': 'Step-by-step guide to reset your email password if you have forgotten it.',
            'keywords': 'password reset, email, forgot password, account recovery',
            'symptoms': 'Cannot log into email account, password not working, account locked',
            'difficulty': 'easy',
            'estimated_time': '15',
            'solution_steps': '1. Go to password reset page\n2. Enter your email address\n3. Check your email for reset link\n4. Create new password',
            'diagnostic_questions': 'Do you have access to your recovery email?',
            'success_rate': '0.95'
        },
        {
            'title': 'Fix Printer Connection Issues',
            'category': 'Hardware',
            'subcategory': 'Printers',
            'content': 'Troubleshooting guide for common printer connectivity problems.',
            'keywords': 'printer, connection, network, troubleshooting, offline',
            'symptoms': 'Printer shows as offline, cannot print, connection error',
            'difficulty': 'medium',
            'estimated_time': '25',
            'solution_steps': '1. Check physical connections\n2. Verify network settings\n3. Restart printer\n4. Test connection',
            'diagnostic_questions': 'Is the printer powered on? Is it connected to the network?',
            'success_rate': '0.85'
        }
    ]
    
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = sample_data[0].keys()
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for row in sample_data:
            writer.writerow(row)
    
    logger.info(f"Sample CSV template created: {file_path}")


def create_sample_excel_template(file_path: str = "sample_articles.xlsx"):
    """Create a sample Excel template for import."""
    workbook = Workbook()
    
    # Create Articles sheet
    articles_sheet = workbook.active
    articles_sheet.title = "Articles"
    
    # Define headers
    headers = [
        'title', 'category', 'subcategory', 'content', 'keywords', 
        'symptoms', 'difficulty', 'estimated_time', 'solution_steps', 
        'diagnostic_questions', 'success_rate'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = articles_sheet.cell(row=1, column=col, value=header.title().replace('_', ' '))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add sample data
    sample_data = [
        [
            'How to Reset Email Password',
            'Email',
            'Password Management',
            'Step-by-step guide to reset your email password if you have forgotten it.',
            'password reset, email, forgot password, account recovery',
            'Cannot log into email account, password not working, account locked',
            'easy',
            '15',
            '1. Go to password reset page\n2. Enter your email address\n3. Check your email for reset link\n4. Create new password',
            'Do you have access to your recovery email?',
            '0.95'
        ],
        [
            'Fix Printer Connection Issues',
            'Hardware',
            'Printers',
            'Troubleshooting guide for common printer connectivity problems.',
            'printer, connection, network, troubleshooting, offline',
            'Printer shows as offline, cannot print, connection error',
            'medium',
            '25',
            '1. Check physical connections\n2. Verify network settings\n3. Restart printer\n4. Test connection',
            'Is the printer powered on? Is it connected to the network?',
            '0.85'
        ]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            articles_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Categories sheet
    categories_sheet = workbook.create_sheet("Categories")
    
    category_headers = ['category', 'subcategory', 'description', 'parent_category']
    for col, header in enumerate(category_headers, 1):
        cell = categories_sheet.cell(row=1, column=col, value=header.title().replace('_', ' '))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    category_data = [
        ['Email', 'Password Management', 'Email password and account management', ''],
        ['Email', 'Connection Issues', 'Email client connectivity problems', ''],
        ['Hardware', 'Printers', 'Printer setup and troubleshooting', ''],
        ['Hardware', 'Network Devices', 'Network equipment and connectivity', ''],
        ['Software', 'Installation', 'Software installation and setup', ''],
        ['Software', 'Updates', 'Software update and maintenance', '']
    ]
    
    for row_idx, row_data in enumerate(category_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            categories_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for sheet in [articles_sheet, categories_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(file_path)
    logger.info(f"Sample Excel template created: {file_path}")


def main():
    """Main function for testing the import system."""
    print("🚀 Knowledge Base Content Import System")
    print("=" * 50)
    
    # Create sample templates
    print("\n📝 Creating sample templates...")
    create_sample_csv_template()
    create_sample_excel_template()
    
    print("\n✅ Sample templates created successfully!")
    print("   - sample_articles.csv")
    print("   - sample_articles.xlsx")
    
    print("\n📚 Usage:")
    print("   from import_system import CSVImporter, JSONImporter, ExcelImporter")
    print("   importer = CSVImporter()")
    print("   result = importer.import_from_csv('sample_articles.csv')")
    
    print("\n🔧 Features:")
    print("   • CSV import with flexible parsing")
    print("   • JSON import with validation")
    print("   • Excel import with multiple sheets")
    print("   • Comprehensive error reporting")
    print("   • Preview mode for validation")
    print("   • Bulk operations support")


if __name__ == "__main__":
    main()
